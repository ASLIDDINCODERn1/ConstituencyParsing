"""
O'zbek tili Dependency Parser — FastAPI backend
Custom Stanza models: Azizwebdev/uzbek-dependency-parser (HuggingFace)
"""

from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from stanza.pipeline.core import DownloadMethod
import stanza
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import logging
import re
import os

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

import threading
_pipeline_lock = threading.Lock()
_preload_done  = False
_preload_err   = None

# ─── UPOS → O'zbek teglari ────────────────────────────────────────────────────

UPOS_UZ = {
    "NOUN":  "OT",       "PROPN": "TO'G'RI OT",
    "VERB":  "FE'L",     "AUX":   "YORD.FE'L",
    "ADJ":   "SIFAT",    "ADV":   "RAVISH",
    "PRON":  "OLMOSH",   "DET":   "ARTIKL",
    "ADP":   "KO'MAKCHI","CCONJ": "BOG'LOVCHI",
    "SCONJ": "BOG'L.(CS)","PART": "YUKL",
    "INTJ":  "UNDOV",    "PUNCT": "TINISH",
    "NUM":   "SON",      "X":     "BOSHQA",
    "SYM":   "BELGI",
}

# ─── Model fayllari ───────────────────────────────────────────────────────────

HF_REPO      = "Azizwebdev/uzbek-dependency-parser"
MODELS_DIR   = os.path.join(os.path.dirname(__file__), "models")

MODEL_FILES = {
    "v8": {
        "tagger":    "uz_v8_tagger.pt",
        "lemmatizer":"uz_v8_lemmatizer.pt",
        "parser":    "uz_v8_parser.pt",
    },
    "v1": {
        "tagger":    "uz_custom_nocharlm_tagger.pt",
        "lemmatizer":"uz_custom_nocharlm_lemmatizer.pt",
        "parser":    "uz_custom_nocharlm_parser.pt",
    },
}
FASTTEXT_FILE = "uz_fasttext.pt"

# ─── O'zbek tokenizer ─────────────────────────────────────────────────────────

_TOKEN_RE = re.compile(
    r"[a-zA-ZʻʼÀ-ÿА-яЁёҚқҲҳҒғÇçĞğİıÖöŞşÜü]"   # word start (Uzbek Latin + Cyrillic)
    r"[a-zA-ZʻʼÀ-ÿА-яЁёҚқҲҳҒғÇçĞğİıÖöŞşÜü0-9''ʻʼ-]*"  # word continuation
    r"|[0-9]+(?:[.,][0-9]+)*"   # numbers with decimals
    r"|[^\s]"                   # any other non-space char (punctuation)
)

def uz_tokenize(text: str) -> list:
    """Simple Uzbek tokenizer — keeps apostrophes inside words (bog'da, o'yna)."""
    return _TOKEN_RE.findall(text.strip())


# ─── Model yuklovchi ──────────────────────────────────────────────────────────

_pipelines: dict = {}   # version → stanza.Pipeline


def download_models(version: str = "v8") -> dict:
    """
    Download custom Uzbek models from HuggingFace Hub.
    Returns local paths dict: {tagger, lemmatizer, parser, fasttext}.
    """
    from huggingface_hub import hf_hub_download

    os.makedirs(MODELS_DIR, exist_ok=True)
    paths = {}

    files = MODEL_FILES.get(version, MODEL_FILES["v8"])
    for role, filename in files.items():
        local = os.path.join(MODELS_DIR, filename)
        if not os.path.exists(local):
            logger.info(f"  ↓ {filename} yuklanmoqda...")
            hf_hub_download(repo_id=HF_REPO, filename=filename, local_dir=MODELS_DIR)
        paths[role] = local

    # FastText embeddings (shared across versions)
    ft_local = os.path.join(MODELS_DIR, FASTTEXT_FILE)
    if not os.path.exists(ft_local):
        logger.info(f"  ↓ {FASTTEXT_FILE} yuklanmoqda (117 MB)...")
        hf_hub_download(repo_id=HF_REPO, filename=FASTTEXT_FILE, local_dir=MODELS_DIR)
    paths["fasttext"] = ft_local

    return paths


def build_pipeline(version: str = "v8") -> stanza.Pipeline:
    """Build a Stanza pipeline with custom Uzbek models."""
    if version in _pipelines:
        return _pipelines[version]

    with _pipeline_lock:
        if version in _pipelines:   # double-check inside lock
            return _pipelines[version]

        logger.info(f"Stanza pipeline qurilmoqda (versiya={version})...")
        paths = download_models(version)

        pipeline = stanza.Pipeline(
            lang="uz",
            processors="tokenize,pos,lemma,depparse",
            tokenize_pretokenized=True,
            allow_unknown_language=True,
            download_method=DownloadMethod.NONE,
            pos_model_path=paths["tagger"],
            pos_pretrain_path=paths["fasttext"],
            lemma_model_path=paths["lemmatizer"],
            depparse_model_path=paths["parser"],
            depparse_pretrain_path=paths["fasttext"],
            use_gpu=False,
            verbose=False,
        )

        _pipelines[version] = pipeline
        logger.info(f"Pipeline ({version}) tayyor!")
        return pipeline


# ─── App ──────────────────────────────────────────────────────────────────────

def _bg_preload():
    """Background thread: download + load v8 pipeline after server starts."""
    global _preload_done, _preload_err
    try:
        build_pipeline("v8")
    except Exception as e:
        _preload_err = str(e)
        logger.warning(f"Preload xato: {e}")
    finally:
        _preload_done = True


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Start model loading in background — uvicorn accepts requests immediately
    threading.Thread(target=_bg_preload, daemon=True).start()
    yield


app = FastAPI(
    title="O'zbek tili Dependency Parser API",
    version="2.0.0",
    description="Custom Stanza models (Azizwebdev/uzbek-dependency-parser)",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Schemalar ────────────────────────────────────────────────────────────────

class ParseRequest(BaseModel):
    text: str
    version: str = "v8"


class TokenData(BaseModel):
    id: int
    text: str
    lemma: str
    upos: str
    upos_uz: str
    head: int
    deprel: str
    is_root: bool
    head_text: Optional[str] = "ROOT"


class ExportRequest(BaseModel):
    sentence: str
    tokens: List[TokenData]


# ─── Endpointlar ──────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    ready = len(_pipelines) > 0
    return {
        "status":         "ok",
        "pipeline_ready": ready,
        "loading":        not _preload_done,
        "error":          _preload_err,
        "loaded_versions": list(_pipelines.keys()),
    }


@app.get("/versions")
def versions():
    return {
        "versions": ["v8", "v1"],
        "default": "v8",
        "descriptions": {
            "v8": "Asosiy model — LAS: 79%, UAS: 85.45% (~3800 gap)",
            "v1": "Boshlang'ich model — LAS: 47% (~500 gap)",
        },
    }


@app.post("/api/parse")
def parse(req: ParseRequest):
    text = req.text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="Matn bo'sh bo'lmasligi kerak")

    version = req.version if req.version in MODEL_FILES else "v8"

    try:
        pipeline = build_pipeline(version)
        tokens_list = uz_tokenize(text)
        if not tokens_list:
            raise HTTPException(status_code=422, detail="Tokenlar aniqlanmadi")

        # Stanza pretokenized input: [[word1, word2, ...]]
        doc = pipeline([tokens_list])

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Parse xato: {e}")
        raise HTTPException(status_code=500, detail=str(e))

    if not doc.sentences:
        raise HTTPException(status_code=422, detail="Gap aniqlanmadi")

    sent = doc.sentences[0]
    word_map = {w.id: w.text for w in sent.words}

    result = []
    for word in sent.words:
        upos    = word.upos or "X"
        head    = word.head if word.head is not None else 0
        is_root = (head == 0)
        result.append({
            "id":       word.id,
            "text":     word.text,
            "lemma":    word.lemma or word.text,
            "upos":     upos,
            "upos_uz":  UPOS_UZ.get(upos, upos),
            "head":     head,
            "deprel":   word.deprel or "dep",
            "is_root":  is_root,
            "head_text":"ROOT" if is_root else word_map.get(head, "?"),
        })

    return {
        "sentence":    text,
        "tokens":      result,
        "version":     version,
        "token_count": len(result),
    }


@app.post("/api/export")
def export_xlsx(req: ExportRequest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dependency Tahlil"

    headers = ["ID", "Token", "Lemma", "UPOS", "UZ Tagi",
               "Head ID", "Head Token", "Deprel", "Root?"]
    ws.append(headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=i)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for t in req.tokens:
        ws.append([
            t.id, t.text, t.lemma, t.upos, t.upos_uz,
            t.head, t.head_text or "ROOT", t.deprel,
            "Ha" if t.is_root else "Yo'q",
        ])

    for col_cells in ws.columns:
        w = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(w + 4, 35)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    fname = (req.sentence[:25] + ".xlsx").replace(" ", "_").replace("/", "-")
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
